import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.utils import timezone
from employees.models import Employee
from training.models import TrainingRecord, TrainingCategory, Provider
from .services import calculate_matrix_summary

def export_training_matrix_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Учет обучения"

    # Стили
    header_font = Font(bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    def apply_header_cell(row, col, value, merge_range=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        cell.fill = light_fill
        if merge_range:
            ws.merge_cells(merge_range)

    # Шапка (1-2 строки)
    apply_header_cell(1, 1, "Фамилия И.О.")
    apply_header_cell(1, 2, "Цех, отдел")
    apply_header_cell(1, 3, "Должность (профессия)")

    # Группы обучения (объединённые ячейки)
    groups = [
        (4, 8, "Программа обучения требованиям охраны труда"),
        (9, 12, "Программа обучения оказанию первой помощи пострадавшим"),
        (13, 16, "Программа обучения по использованию (применению) СИЗ"),
    ]
    for start, end, title in groups:
        apply_header_cell(1, start, title, f"{openpyxl.utils.get_column_letter(start)}1:{openpyxl.utils.get_column_letter(end)}1")

    # Подзаголовки
    subheaders = [
        "Подлежит обучению по охране труда", "Работодатель программа Б", 
        "Учебный центр Программа А и Б", "Дата обучения",
        "Освобожден от прохождения обучения по охране труда",
        "Не требуется прохождение стажировки на рабочем месте",
        "Освобожден от прохождения первичного инструктажа по охране труда",
        "Подлежит обучению по оказанию первой помощи пострадавшим",
        "Работодатель", "Учебный центр", "Дата обучения",
        "Подлежит обучению по использованию (применению) СИЗ",
        "Работодатель", "Учебный центр", "Дата обучения"
    ]
    for col, title in enumerate(subheaders, 4):
        apply_header_cell(2, col, title)

    # Данные
    employees = Employee.objects.filter(is_active=True).select_related("position__department").prefetch_related("training_records")
    row_num = 3
    for emp in employees:
        ws.cell(row=row_num, column=1, value=f"{emp.last_name} {emp.first_name[0]}.").border = thin_border
        ws.cell(row=row_num, column=2, value=emp.position.department.name).border = thin_border
        ws.cell(row_num, column=3, value=emp.position.name).border = thin_border

        records = {r.category: r for r in emp.training_records.all()}

        # Логика заполнения колонок 4-18
        # Колонка 4: Подлежит обучению по ОТ (1 если должность требует)
        # Для упрощения: 1 если не exempt_ot_training
        ot_req = 0 if emp.exempt_ot_training else 1
        ws.cell(row=row_num, column=4, value=ot_req if ot_req else "").border = thin_border
        
        ot_rec = records.get(TrainingCategory.OT)
        if ot_rec:
            ws.cell(row=row_num, column=5 if ot_rec.provider == Provider.EMPLOYER else 6, value=1).border = thin_border
            ws.cell(row=row_num, column=7, value=ot_rec.training_date.strftime("%m/%d/%y")).border = thin_border

        ws.cell(row=row_num, column=8, value=1 if emp.exempt_ot_training else "").border = thin_border
        ws.cell(row=row_num, column=9, value=1 if emp.exempt_internship else "").border = thin_border
        ws.cell(row=row_num, column=10, value=1 if emp.exempt_primary_briefing else "").border = thin_border

        # ППП
        ppp_req = 1
        ws.cell(row=row_num, column=11, value=ppp_req).border = thin_border
        ppp_rec = records.get(TrainingCategory.FIRST_AID)
        if ppp_rec:
            ws.cell(row=row_num, column=12 if ppp_rec.provider == Provider.EMPLOYER else 13, value=1).border = thin_border
            ws.cell(row=row_num, column=14, value=ppp_rec.training_date.strftime("%m/%d/%y")).border = thin_border

        # СИЗ
        siz_req = 1
        ws.cell(row=row_num, column=15, value=siz_req).border = thin_border
        siz_rec = records.get(TrainingCategory.PPE)
        if siz_rec:
            ws.cell(row=row_num, column=16 if siz_rec.provider == Provider.EMPLOYER else 17, value=1).border = thin_border
            ws.cell(row=row_num, column=18, value=siz_rec.training_date.strftime("%m/%d/%y")).border = thin_border

        row_num += 1

    # Итоги
    summary = calculate_matrix_summary()
    summary_rows = [
        ("Общее количество работников, подлежащих обучению по охране труда", summary["total_ot"]),
        ("Общее количество работников, освобожденных от прохождения обучения по охране труда", summary["exempt_ot"]),
        ("Общее количество работников, которым не требуется прохождение стажировки на рабочем месте", summary["no_internship"]),
        ("Общее количество работников, освобожденных от прохождения первичного инструктажа по охране труда", summary["exempt_briefing"]),
        ("Общее количество работников, подлежащих обучению ППП", summary["ppp"]),
        ("Общее количество работников, подлежащих обучению СИЗ", summary["siz"]),
    ]
    row_num += 2
    for title, val in summary_rows:
        ws.cell(row=row_num, column=1, value=title).border = thin_border
        ws.cell(row=row_num, column=2, value=title).border = thin_border
        ws.cell(row=row_num, column=3, value=val).border = thin_border
        row_num += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    return wb